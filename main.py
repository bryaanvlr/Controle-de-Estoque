from openpyxl import load_workbook
from openpyxl.styles import Alignment
from copy import copy

arquivo = load_workbook("projeto_controle_estoque_150_produtos.xlsx")
relatorio = arquivo.create_sheet("Relatório")
aba = arquivo.active

valor_total_estoque = 0
maior_estoque = 0
produto_maior_estoque = ""
menor_estoque = 99999
produto_menor_estoque = ""
estoque_baixo = 0


for linha in range(2, aba.max_row + 1):
    produto = aba.cell(row=linha, column=1).value
    categoria = aba.cell(row=linha, column=2).value
    estoque = aba.cell(row=linha, column=3).value
    preco = aba.cell(row=linha, column=4).value

    quantidade_produto = aba.max_row - 1

    valor_estoque = estoque * preco
    valor_total_estoque += valor_estoque

    if estoque > maior_estoque:
        maior_estoque = estoque
        produto_maior_estoque = produto
    if estoque < menor_estoque:
        menor_estoque = estoque
        produto_menor_estoque = produto
    if estoque < 10:
        estoque_baixo += 1

valor_formatado = f"{valor_total_estoque:,.2f}"
valor_formatado = valor_formatado.replace(",", "X").replace(".", ",").replace("X", ".")

print(f"R$ {valor_formatado}")
print(estoque_baixo)


relatorio["A1"] = "Total de produtos cadastrados"
relatorio["A2"] = "Valor total do estoque"
relatorio["A3"] = "Produto com maior estoque"
relatorio["A4"] = "Quantidade do maior estoque"
relatorio["A5"] = "Produto com menor estoque"
relatorio["A6"] = "Quantidade do menor estoque"
relatorio["A7"] = "Produtos com estoque baixo (menos de 10 unidades)"

relatorio["B1"] = quantidade_produto
relatorio["B2"] = f"R$ {valor_formatado}"
relatorio["B3"] = produto_maior_estoque
relatorio["B4"] = maior_estoque
relatorio["B5"] = produto_menor_estoque
relatorio["B6"] = menor_estoque
relatorio["B7"] = estoque_baixo

for linha in range(1, 8):
    
    for coluna in range(1, 3):
        celula_relatorio = relatorio.cell(row=linha, column=coluna)
        celula_modelo = aba.cell(row=2, column=1)
        
        celula_relatorio.font = copy(celula_modelo.font)
        celula_relatorio.fill = copy(celula_modelo.fill)
        celula_relatorio.border = copy(celula_modelo.border)
        celula_relatorio.alignment = copy(celula_modelo.alignment)
    for linha in range(1, 8):
        relatorio.cell(row=linha, column=2).alignment = Alignment(horizontal="left")

        relatorio.column_dimensions["A"].width = 45
        relatorio.column_dimensions["B"].width = 30

arquivo.save("projeto_controle_estoque_150_produtos2.xlsx")